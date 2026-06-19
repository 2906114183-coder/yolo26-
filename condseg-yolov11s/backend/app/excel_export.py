from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "图片名",
    "缺陷编号",
    "缺陷类别",
    "置信度",
    "左上角X",
    "左上角Y",
    "右下角X",
    "右下角Y",
    "人工状态",
    "备注",
    "最终结论",
]


def export_job_to_excel(job: dict, output_path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "检测明细"
    sheet.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="E8F0EA")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for image in job.get("images", []):
        active_detections = [det for det in image.get("detections", []) if not det.get("deleted")]
        if not active_detections:
            sheet.append(
                [
                    image.get("original_name", ""),
                    "",
                    "未检测到缺陷",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    image.get("remark", ""),
                    "合格",
                ]
            )
            continue

        for index, det in enumerate(active_detections, start=1):
            x1, y1, x2, y2 = det.get("bbox", ["", "", "", ""])
            status = det.get("status", "unconfirmed")
            conclusion = "合格" if status == "false_positive" else "待复核" if status == "unconfirmed" else "不合格"
            sheet.append(
                [
                    image.get("original_name", ""),
                    index,
                    det.get("class_name", ""),
                    det.get("confidence", ""),
                    x1,
                    y1,
                    x2,
                    y2,
                    status,
                    det.get("remark", ""),
                    conclusion,
                ]
            )

    for column_index, column_cells in enumerate(sheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 10), 32)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
